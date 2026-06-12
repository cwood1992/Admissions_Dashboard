"""Build the cash-flow projection data for the unified dashboard's Cash tab.

Joins three sources:
  1. DISBURSMNT 26-27.xlsx  - per-class disbursement schedule (Title IV 1st/2nd
     dates on the class row, VA release dates on the companion row below it).
  2. "NNN Class.xlsx" expected-funds workbooks (one per class) - per-student
     expected funding by source. PII never leaves this script: only counts and
     per-class/program sums are written out (FERPA).
  3. ledger/cohort_ledger.json - start projections for cohorts that do not yet
     have an expected-funds workbook.

Timing model (confirmed with FA):
  - Title IV (PELL/SEOG/SUB/UNSUB/PLUS): 50% at 1st disb date, 50% at 2nd,
    net of origination fees (1.057% SUB/UNSUB, 4.228% PLUS).
  - VA: 100% on the first VA release date. SCHOL: at 1st disb date.
  - Cash payers (no expected funds): counted, no dollars modeled.

Run from repo root:  python ledger\\build_cashflow.py
"""

from __future__ import annotations

import json
import os
import re
import shutil
import sys
import tempfile
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# ---------------------------------------------------------------- CONFIG ----
FA_DIR = Path(r"C:\Users\Clanton\The Ocean Corporation"
              r"\TOC-FinancialAid - Documents\Expected Funds\From 561 Class and above")
SCHEDULE_PATH = Path(r"C:\Users\Clanton\The Ocean Corporation"
                     r"\TOC-FinancialAid - Documents\Disbursement Dates\DISBURSMNT 26-27.xlsx")
SCHEDULE_SHEET = "24-26"
LEDGER_PATH = ROOT / "ledger" / "cohort_ledger.json"
OUT_JSON = ROOT / "cash" / "data" / "cashflow.json"
OUT_JS = ROOT / "cash" / "data" / "cashflow.js"

FUNDS = ["VA", "PELL", "SEOG", "SUB", "UNSUB", "PLUS", "SCHOL"]
TITLE_IV = ["PELL", "SEOG", "SUB", "UNSUB", "PLUS"]  # split 50/50 disb1/disb2
ORIGINATION_FEES = {"SUB": 0.01057, "UNSUB": 0.01057, "PLUS": 0.04228}

ASSUMPTIONS = [
    "Title IV (Pell/SEOG/Sub/Unsub/PLUS) arrives 50% on the 1st disbursement "
    "date and 50% on the 2nd, net of origination fees (1.057% Sub/Unsub, "
    "4.228% PLUS).",
    "VA benefits arrive 100% on the first VA release date (~46th class day).",
    "Scholarships are modeled on the 1st disbursement date.",
    "Cash payers are counted but no dollars are modeled for them - their "
    "payment timing is not in the expected-funds files.",
    "Re-entry students are included in their class totals on that class's "
    "schedule dates (actual request dates may differ).",
    "Title IV returns/refunds (R2T4) are not modeled.",
    "Future cohorts without an expected-funds file use the per-program average "
    "net funding per student from filed classes x projected starts "
    "(low/mid/high) from the cohort ledger.",
    "Classes beyond the disbursement schedule use dates extrapolated from the "
    "average calendar offsets of scheduled classes.",
]


def _tmp_copy(path: Path) -> Path:
    """Copy a OneDrive file to %TEMP% before opening.

    The source files are Files-On-Demand placeholders and are often locked by
    Excel; reading the original directly raises PermissionError.
    """
    dst = Path(tempfile.gettempdir()) / f"cashflow_{path.name}"
    try:
        shutil.copyfile(path, dst)
    except PermissionError:
        # Python's open() can't hydrate a cloud-only placeholder, but
        # PowerShell's Copy-Item can.
        import subprocess
        subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             f'Copy-Item -LiteralPath "{path}" -Destination "{dst}" -Force'],
            check=True, capture_output=True)
    return dst


def _num(value) -> float:
    """Coerce a cell to a dollar amount; text like '-', 'Declined', 'HOLD' -> 0."""
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _iso(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


# ------------------------------------------------------------- schedule -----
def parse_schedule() -> dict[int, dict]:
    """Sheet '24-26': class row (start, disb1, disb2, grad) followed by a
    companion row carrying the two VA release dates in the same columns."""
    wb = openpyxl.load_workbook(_tmp_copy(SCHEDULE_PATH), data_only=True)
    ws = wb[SCHEDULE_SHEET]
    schedule: dict[int, dict] = {}
    current: dict | None = None
    for row in ws.iter_rows(values_only=True):
        a, start, d1, d2, grad = (list(row) + [None] * 5)[:5]
        if isinstance(a, (int, float)) and 400 < a < 700:
            current = {
                "start": _iso(start), "disb1": _iso(d1), "disb2": _iso(d2),
                "grad": _iso(grad), "va1": None, "va2": None, "derived": False,
            }
            schedule[int(a)] = current
        elif current is not None and a is None and (_iso(d1) or _iso(d2)):
            current["va1"], current["va2"] = _iso(d1), _iso(d2)
            current = None
    return schedule


def extrapolate_schedule(schedule: dict[int, dict], class_num: int,
                         start_iso: str) -> dict:
    """Derive dates for a class beyond the schedule from mean calendar offsets."""
    offsets: dict[str, list[int]] = {k: [] for k in ("disb1", "disb2", "va1", "va2", "grad")}
    for rec in schedule.values():
        if rec.get("derived") or not rec.get("start"):
            continue
        start = date.fromisoformat(rec["start"])
        for key in offsets:
            if rec.get(key):
                d = date.fromisoformat(rec[key])
                if d > start:  # guards the class-560 typo year (1925)
                    offsets[key].append((d - start).days)
    start = date.fromisoformat(start_iso)
    rec = {"start": start_iso, "derived": True}
    for key, days in offsets.items():
        rec[key] = (start + timedelta(days=round(sum(days) / len(days)))).isoformat() if days else None
    return rec


# ------------------------------------------------------- expected funds -----
HEADER_FUND_MAP = [
    (re.compile(r"^VA \$"), "VA"),
    (re.compile(r"^PELL"), "PELL"),
    (re.compile(r"^SEOG"), "SEOG"),
    (re.compile(r"^SUB\b"), "SUB"),
    (re.compile(r"^UNSUB"), "UNSUB"),
    (re.compile(r"^PLUS"), "PLUS"),
    (re.compile(r"^SCHOL"), "SCHOL"),
    (re.compile(r"^SUM$"), "SUM"),
]


def _colmap(header_row: tuple) -> dict[int, str] | None:
    """Map column index -> fund key for a section header row."""
    labels = {i: str(v).strip().upper() for i, v in enumerate(header_row)
              if isinstance(v, str) and v.strip()}
    if "PROGRAM" not in labels.values():
        return None
    mapping: dict[int, str] = {}
    for i, label in labels.items():
        for pattern, fund in HEADER_FUND_MAP:
            if pattern.match(label):
                mapping[i] = fund
                break
    return mapping if any(f in mapping.values() for f in FUNDS) else None


PROGRAM_RE = re.compile(r"(UDT|NDT|U|N)\s*(\d{3})\s*(NC)?", re.IGNORECASE)


def _parse_program(label: str, file_class: int) -> tuple[int, str] | None:
    """Normalize a PROGRAM cell to (class_num, program_group)."""
    text = label.strip()
    if not text or text.upper() == "PROGRAM":
        return None
    if re.search(r"RE-?ENTRY", text, re.IGNORECASE):
        m = re.search(r"(\d{3})", text)
        return (int(m.group(1)) if m else file_class, "Re-entry")
    m = PROGRAM_RE.search(text)
    if m:
        base = "UDT" if m.group(1).upper().startswith("U") else "NDT"
        program = f"{base}-NC" if m.group(3) else base
        return int(m.group(2)), program
    if text.upper() in ("UDT", "NDT"):  # 561-era labels without class number
        return file_class, text.upper()
    return None


def parse_expected_funds(path: Path, file_class: int) -> dict:
    """First worksheet only (the live 'Check List'); other sheets are drafts.

    Returns {(class_num, program): {students, cash_payers, funds{...gross}}}
    plus a reconciliation list of (row_sum_col, row_parsed_total) pairs.
    """
    wb = openpyxl.load_workbook(_tmp_copy(path), data_only=True)
    ws = wb.worksheets[0]
    sections: dict[tuple[int, str], dict] = {}
    recon = {"sum_col": 0.0, "parsed": 0.0}
    colmap: dict[int, str] | None = None
    for row in ws.iter_rows(values_only=True):
        new_map = _colmap(row)
        if new_map:
            colmap = new_map
            continue
        if not colmap or not isinstance(row[0], str):
            continue
        parsed = _parse_program(row[0], file_class)
        if parsed is None:
            continue
        key = parsed
        amounts = {f: 0.0 for f in FUNDS}
        sum_col = 0.0
        for i, fund in colmap.items():
            if i >= len(row):
                continue
            if fund == "SUM":
                sum_col += _num(row[i])
            elif fund in amounts:
                amounts[fund] += _num(row[i])
        total = sum(amounts.values())
        if total == 0 and sum_col == 0 and all(
                v is None or (isinstance(v, str) and not v.strip())
                for j, v in enumerate(row) if j in colmap):
            # blank styled row, not a student
            continue
        rec = sections.setdefault(key, {
            "students": 0, "cash_payers": 0, "funds": {f: 0.0 for f in FUNDS}})
        rec["students"] += 1
        if total <= 0:
            rec["cash_payers"] += 1
        for f in FUNDS:
            rec["funds"][f] += amounts[f]
        recon["sum_col"] += sum_col
        recon["parsed"] += total
    return {"sections": sections, "recon": recon}


# --------------------------------------------------------------- events -----
def net_amount(fund: str, gross: float) -> float:
    return round(gross * (1.0 - ORIGINATION_FEES.get(fund, 0.0)), 2)


def fund_events(class_num: int, program: str, dates: dict, funds: dict,
                kind: str, cohort_label: str, bands: dict | None = None):
    """Build event dicts for one (class, program) funding block.

    funds holds gross mid-case dollars per fund; bands optionally holds gross
    low/high dollars per fund: {'low': {fund: gross}, 'high': {...}}.
    Each (fund, tranche) pair becomes one event; Title IV splits 50/50 across
    disb1/disb2, VA lands whole on va1, SCHOL on disb1.
    """
    tranches = []  # (fund, tranche_key, fraction_of_gross)
    for fund in TITLE_IV:
        tranches.append((fund, "disb1", 0.5))
        tranches.append((fund, "disb2", 0.5))
    tranches.append(("VA", "va1", 1.0))
    tranches.append(("SCHOL", "disb1", 1.0))

    events = []
    for fund, tranche, fraction in tranches:
        when = dates.get(tranche)
        if not when:
            continue
        ev = {
            "date": when, "class": class_num, "cohort": cohort_label,
            "program": program, "fund": fund, "tranche": tranche, "kind": kind,
            "amount": net_amount(fund, funds.get(fund, 0.0) * fraction),
        }
        if bands:
            ev["low"] = net_amount(fund, bands["low"].get(fund, 0.0) * fraction)
            ev["high"] = net_amount(fund, bands["high"].get(fund, 0.0) * fraction)
        if ev["amount"] == 0 and ev.get("high", 0) == 0:
            continue
        events.append(ev)
    return events


# ----------------------------------------------------------------- main -----
def main() -> int:
    schedule = parse_schedule()

    # 1. parse all expected-funds workbooks --------------------------------
    filed: dict[int, dict] = {}  # class -> {(class,prog): rec}
    file_meta = []
    for path in sorted(FA_DIR.glob("* Class.xlsx")):
        if path.name.startswith("~$"):
            continue
        m = re.match(r"(\d{3}) Class", path.stem)
        if not m:
            continue
        file_class = int(m.group(1))
        parsed = parse_expected_funds(path, file_class)
        filed[file_class] = parsed["sections"]
        mtime = datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()
        file_meta.append({"class": file_class, "file": path.name, "modified": mtime})
        r = parsed["recon"]
        flag = "OK" if abs(r["sum_col"] - r["parsed"]) < 1 else "MISMATCH"
        print(f"  {path.name}: parsed ${r['parsed']:,.0f} vs SUM col "
              f"${r['sum_col']:,.0f}  [{flag}]")

    filed_classes = set(filed)

    # 2. per-program average gross funding per student (for projections) ---
    avg: dict[str, dict] = {}
    for sections in filed.values():
        for (cls, program), rec in sections.items():
            if program == "Re-entry":
                continue
            a = avg.setdefault(program, {"students": 0, "funds": {f: 0.0 for f in FUNDS}})
            a["students"] += rec["students"]
            for f in FUNDS:
                a["funds"][f] += rec["funds"][f]
    per_student = {
        prog: {f: (a["funds"][f] / a["students"] if a["students"] else 0.0) for f in FUNDS}
        for prog, a in avg.items()
    }

    # 3. expected events from filed classes --------------------------------
    events: list[dict] = []
    class_summaries: dict[int, dict] = {}
    for file_class, sections in filed.items():
        for (cls, program), rec in sections.items():
            dates = schedule.get(cls)
            if not dates:
                print(f"  WARNING: class {cls} has no schedule row - skipped")
                continue
            label = f"{program} {cls}"
            events.extend(fund_events(cls, program, dates, rec["funds"],
                                      "expected", label))
            summary = class_summaries.setdefault(cls, {
                "class": cls, "source": "file", "dates": dates, "programs": []})
            summary["programs"].append({
                "program": program, "students": rec["students"],
                "cash_payers": rec["cash_payers"],
                "funds_gross": {f: round(rec["funds"][f], 2) for f in FUNDS},
                "total_net": round(sum(
                    net_amount(f, rec["funds"][f]) for f in FUNDS), 2),
            })

    # 4. projected events from the cohort ledger ---------------------------
    ledger = json.loads(LEDGER_PATH.read_text(encoding="utf-8"))
    for cohort in ledger["cohorts"]:
        if cohort.get("proj_mid") is None:
            continue
        m = re.match(r"(UDT|NDT)(\d{3})(NC)?$", cohort["cohort"])
        if not m:
            continue
        cls = int(m.group(2))
        if cls in filed_classes:
            continue  # real expected-funds data wins
        program = ("UDT" if m.group(1) == "UDT" else "NDT") + ("-NC" if m.group(3) else "")
        mix = per_student.get(program)
        if not mix:
            print(f"  WARNING: no funding mix for program {program} - skipped {cohort['cohort']}")
            continue
        dates = schedule.get(cls)
        if not dates:
            if not cohort.get("start_date"):
                print(f"  WARNING: {cohort['cohort']} has no schedule and no start date - skipped")
                continue
            dates = extrapolate_schedule(schedule, cls, cohort["start_date"])
            schedule[cls] = dates
        gross = {f: mix[f] * cohort["proj_mid"] for f in FUNDS}
        bands = {
            "low": {f: mix[f] * cohort["proj_low"] for f in FUNDS},
            "high": {f: mix[f] * cohort["proj_high"] for f in FUNDS},
        }
        label = f"{program} {cls}"
        events.extend(fund_events(cls, program, dates, gross, "projected",
                                  label, bands))
        summary = class_summaries.setdefault(cls, {
            "class": cls,
            "source": "projected-derived-dates" if dates.get("derived") else "projected",
            "dates": dates, "programs": []})
        summary["programs"].append({
            "program": program,
            "proj_starts": {"low": cohort["proj_low"], "mid": cohort["proj_mid"],
                            "high": cohort["proj_high"]},
            "funds_gross": {f: round(gross[f], 2) for f in FUNDS},
            "total_net": round(sum(net_amount(f, gross[f]) for f in FUNDS), 2),
        })

    events.sort(key=lambda e: (e["date"], e["class"], e["fund"]))

    # 5. write outputs ------------------------------------------------------
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "ledger_snapshot_date": ledger.get("snapshot_date"),
        "assumptions": ASSUMPTIONS,
        "sources": {
            "schedule_file": SCHEDULE_PATH.name,
            "expected_funds_files": file_meta,
        },
        "funds": FUNDS,
        "per_student_avg_gross": {
            prog: {f: round(v, 2) for f, v in mix.items()}
            for prog, mix in per_student.items()
        },
        "classes": [class_summaries[c] for c in sorted(class_summaries)],
        "events": events,
    }
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    text = json.dumps(payload, indent=1)
    OUT_JSON.write_text(text, encoding="utf-8")
    OUT_JS.write_text("window.CASHFLOW = " + text + ";\n", encoding="utf-8")

    n_exp = sum(1 for e in events if e["kind"] == "expected")
    total_exp = sum(e["amount"] for e in events if e["kind"] == "expected")
    total_proj = sum(e["amount"] for e in events if e["kind"] == "projected")
    print(f"\n  {len(events)} events ({n_exp} expected / {len(events) - n_exp} projected)")
    print(f"  expected net inflows:  ${total_exp:,.0f}")
    print(f"  projected net (mid):   ${total_proj:,.0f}")
    print(f"  wrote {OUT_JSON.relative_to(ROOT)} and {OUT_JS.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
